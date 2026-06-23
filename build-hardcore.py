"""
build-hardcore.py
-----------------
Builds data/hardcore.json from the "Restricted_Hardcore Mode Info & Hardcore
Bosses" spreadsheet + the repo's trainers.json/species.json.

The spreadsheet supplies STRUCTURE (which bosses, what category, what order, the
rules text). The authoritative team data already lives in trainers.json's
`hardcore` field, so we only need to bind each spreadsheet boss to a trainer ID.
We do that by matching the boss's team species (cols 4/9/14/19/24/29) against
each trainer's hardcore team, with the boss's proper name as a strong signal.
This disambiguates duplicates like first-battle Brock (Lv13-16) vs rematch Brock.

    python build-hardcore.py

Re-run if the spreadsheet or trainers.json changes. Reports any unmatched boss.
"""
import glob, json, os, re

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(ROOT, 'data')

import openpyxl

SHEETS = ['Kanto Leaders', 'Kanto Rematch', 'Johto Leaders', 'Rivals',
          'Team Rocket', 'Mini Bosses', 'Optional Bosses', 'Indigo League', 'Postgame']
SPECIES_COLS = [4, 9, 14, 19, 24, 29]
NAME_COL = 2

ALIASES = {  # spreadsheet shorthand -> token used in species key
    'bm': 'bloodmoon', 'galar': 'galarian', 'alola': 'alolan',
    'hisui': 'hisuian', 'paldea': 'paldean', 'sevii': 'sevii',
}

def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def load_json(name):
    with open(os.path.join(DATA, name), encoding='utf-8') as f:
        return json.load(f)

def build_species_index(species):
    idx = {}
    for s in species.values():
        for label in (s.get('name'), s.get('key')):
            if label:
                idx.setdefault(norm(label), s['ID'])
    return idx

def resolve_species(name, idx):
    n = norm(name)
    if n in idx:
        return idx[n]
    # try expanding a trailing form shorthand: "Ursaluna BM" / "Corsola-Galar"
    parts = re.split(r'[\s\-]+', str(name).strip())
    if len(parts) >= 2:
        base, suf = parts[0], parts[-1].lower()
        suf = ALIASES.get(suf, suf)
        cand = norm(base + suf)
        if cand in idx:
            return idx[cand]
        if norm(base) in idx:           # fall back to base species (good enough for ID matching)
            return idx[norm(base)]
    return None

def get(row, i):
    return row[i] if i < len(row) and row[i] is not None else None

def extract_bosses(ws, species_idx, ws_f=None):
    bosses = []
    for ridx, row in enumerate(ws.iter_rows(values_only=True)):
        name = get(row, NAME_COL)
        if not name or not isinstance(name, str):
            continue
        nm = ' '.join(str(name).split())
        if not nm or 'BATTLE EFFECT' in nm.upper() or 'BASE STAT' in nm.upper():
            continue
        sp_names = [get(row, c) for c in SPECIES_COLS]
        sp_names = [s for s in sp_names if s and isinstance(s, str)]
        ids = [resolve_species(s, species_idx) for s in sp_names]
        ids = [i for i in ids if i]
        if len(ids) < 2:                # real boss row: species cols resolve to species
            continue
        # trainer sprite = the =IMAGE("url") in the cell directly above the name (same column)
        sprite = ''
        if ws_f is not None and ridx >= 1:
            fv = ws_f.cell(row=ridx, column=NAME_COL + 1).value  # name is iter row ridx -> 1-indexed ridx+1; cell above = row ridx
            if isinstance(fv, str):
                m = re.search(r'IMAGE\("([^"]+)"', fv)
                if m:
                    sprite = m.group(1)
        bosses.append({'name': nm, 'species_names': sp_names, 'species_ids': ids, 'sprite': sprite, 'row': ridx + 1})
    return bosses

def proper_name(boss_name):
    # last word is usually the trainer's proper name (BROCK, BRENDAN, ARCHER)
    words = re.findall(r"[A-Za-z.'&]+", boss_name)
    return words[-1].lower() if words else ''

def match_trainer(boss, trainers):
    target = set(boss['species_ids'])
    pname = proper_name(boss['name'])
    best, best_score = None, -1
    for t in trainers.values():
        team = t.get('hardcore') or []
        if not team:
            continue
        team_ids = set(m['species'] for m in team)
        overlap = len(target & team_ids)
        name_hit = 1 if pname and pname in t['name'].lower() else 0
        # require some species overlap; name match is a strong tiebreak
        score = overlap + 4 * name_hit
        # prefer same team size
        score -= abs(len(team) - len(boss['species_ids'])) * 0.1
        if score > best_score:
            best, best_score = t, score
    # accept only confident matches
    if best and (len(target & set(m['species'] for m in best['hardcore'])) >= 2
                 or (proper_name(boss['name']) and proper_name(boss['name']) in best['name'].lower())):
        return best['ID'], best_score
    return None, best_score

def to_num(cap):
    nums = re.findall(r'\d+', str(cap))
    return max(int(n) for n in nums) if nums else None

def clean_cap(v):
    s = str(v).strip()
    return s[:-2] if s.endswith('.0') else s

GENERIC = set(('lass ace youngster catcher bug super nerd twins beauty channeler leader trainer '
               'pkmn gym elite four rocket admin grunt guard sis bro left right swimmer fisherman '
               'gentleman picnicker camper hiker biker sailor rocker burglar tamer cooltrainer '
               'blackbelt juggler psychic scientist').split())

def match_order_trainer(name, capnum, trainers):
    # strip parentheticals like "(REMATCH)" and split into words
    words = set(re.findall(r"[a-z']+", re.sub(r'\(.*?\)', ' ', name).lower())) - {'rematch'}
    proper = words - GENERIC                  # the identifying name(s), e.g. {'anne'}, {'brock'}
    if not words:
        return None
    best = None                               # key = (-overlap, levelDiff)
    for t in trainers.values():
        toks = set(re.findall(r"[a-z']+", t['name'].lower()))
        overlap = words & toks
        # require a proper-name hit, or a 2+ word overlap — avoids matching on class alone
        if not (proper & toks) and len(overlap) < 2:
            continue
        team = t.get('hardcore') or []
        if not team:
            continue
        mx = max(m['level'] for m in team)
        diff = abs(mx - (capnum if capnum is not None else mx))
        key = (-len(overlap), diff)
        if best is None or key < best[0]:
            best = (key, t['ID'])
    if best:
        return best[1]
    # Stale grunts/guards -> nearest Team Rocket Grunt by resolved team level.
    low = name.lower()
    if 'grunt' in low or 'guard' in low:
        sc = {101: -2, 102: -1, 103: 0, 104: -3}
        rl = lambda c: (100 + sc[c]) if c in sc else (100 if c > 100 else c)
        cand = None
        for t in trainers.values():
            if 'team rocket grunt' not in t['name'].lower():
                continue
            team = t.get('hardcore') or []
            if not team:
                continue
            mx = max(rl(m['level']) for m in team)
            d = abs(mx - (capnum if capnum is not None else mx))
            if cand is None or d < cand[0]:
                cand = (d, t['ID'])
        if cand:
            return cand[1]
    return None

def resolve_link(loc, loc_map):
    # loc is an internal ref like "Rivals!C5" or "'Mini Bosses'!C3" pointing at a boss block.
    if not loc or '!' not in loc:
        return None
    sheet, cell = loc.rsplit('!', 1)
    sheet = sheet.strip().strip("'")
    m = re.search(r'(\d+)', cell)
    if not m:
        return None
    L = int(m.group(1))
    lst = loc_map.get(sheet)
    if not lst:
        return None
    ge = [(r, t) for (r, t) in lst if r >= L]   # link targets the block top; name row is at/after L
    return (min(ge)[1] if ge else min(lst, key=lambda rt: abs(rt[0] - L))[1])

def parse_trainer_order(ws, trainers, loc_map):
    entries, optional, expecting_loc = [], False, False
    for row in ws.iter_rows():
        cv = lambda i: row[i].value if i < len(row) and row[i].value is not None else None
        c2, c3, c5 = cv(2), cv(3), cv(5)
        if c2 and 'OPTIONAL' in str(c2).upper():
            optional = True
            continue
        name = ' '.join(str(c3).split()) if c3 else ''
        link = row[3].hyperlink.location if (len(row) > 3 and row[3].hyperlink) else None
        if c3 and c5 is not None:
            entries.append({'name': name, 'cap': clean_cap(c5), 'optional': optional, 'location': None, 'link': link})
            optional, expecting_loc = False, True
        elif c3 and expecting_loc:
            entries[-1]['location'] = name
            expecting_loc = False
        elif c3:  # name row missing its cap cell -> new entry, inherit prior cap
            entries.append({'name': name, 'cap': entries[-1]['cap'] if entries else '',
                            'optional': optional, 'location': None, 'link': link})
            optional, expecting_loc = False, True
    for e in entries:
        tid = resolve_link(e.pop('link', None), loc_map)   # exact mapping from the sheet's hyperlink
        e['trainerId'] = tid if tid else match_order_trainer(e['name'], to_num(e['cap']), trainers)
    return entries

INFO_HEADERS = {
    'HARDCORE GENERAL RESTRICTIONS AND CHANGES': ('General Restrictions', 'list'),
    'LEVEL CAPS': ('Level Caps', 'caps'),
    'HOW TO GET FOCUS SASH AND CHOICE ITEMS FROM QUESTS': ('Focus Sash & Choice Items', 'list'),
    'ABOUT THE SHEET': ('About This Data', 'list'),
    'ABILITY CHANGES': ('Ability Changes', 'abilities'),
}

def build_info(ws):
    rows = list(ws.iter_rows(values_only=True))
    def g(r, c):
        return rows[r][c] if c < len(rows[r]) and rows[r][c] is not None else None
    cell = lambda v: ' '.join(str(v).split()) if v is not None else ''

    sections, cur, ability_row = [], None, None
    for i in range(len(rows)):
        h = cell(g(i, 2))
        if h:
            if h in INFO_HEADERS:
                title, typ = INFO_HEADERS[h]
                cur = {'title': title, 'type': typ, 'items': []}
                sections.append(cur)
                if typ == 'abilities':
                    ability_row, cur = i, None
                continue
            cur = None  # unrecognized header ends the current list
        if cur and cur['type'] in ('list', 'caps'):
            v = cell(g(i, 4))
            if v:
                cur['items'].append(v)

    for s in sections:
        if s['type'] == 'caps':
            out = []
            for it in s['items']:
                m = re.search(r'\((\d+)\)\s*$', it)
                out.append({'label': re.sub(r'\s*\(\d+\)\s*$', '', it), 'cap': int(m.group(1)) if m else None})
            s['items'] = out

    ab = next((s for s in sections if s['type'] == 'abilities'), None)
    if ab and ability_row is not None:
        pairs = []
        for ban_c, chg_c in [(4, 9), (16, 21), (30, 35)]:
            grp = []
            for i in range(ability_row + 1, min(ability_row + 40, len(rows))):
                ban, chg = cell(g(i, ban_c)), cell(g(i, chg_c))
                if ban.upper() in ('BANNED MOVES', 'PHYSICAL'):
                    break
                if ban and chg and ban.upper() != 'BANNED':
                    grp.append(ban + ' → ' + chg)
                elif ban and not chg and not ban.startswith('('):
                    if 'KEEP' in ban.upper():
                        grp.append(ban)
                    elif grp and '→' in grp[-1]:           # also-banned, shares prev change
                        left, right = grp[-1].rsplit('→', 1)
                        grp[-1] = left.strip() + ', ' + ban + ' → ' + right.strip()
            pairs.extend(grp)
        ab['items'] = pairs

    return [s for s in sections if s['items']]

def main():
    f = glob.glob(os.path.join('C:/Users/anten/Downloads', '*Hardcore*Radical Red*.xlsx'))[0]
    print('reading', os.path.basename(f).encode('ascii', 'replace').decode())
    wb = openpyxl.load_workbook(f, data_only=True)
    wb_f = openpyxl.load_workbook(f, data_only=False)  # formulas, for the =IMAGE() sprite URLs

    species = load_json('species.json')
    trainers = load_json('trainers.json')
    sidx = build_species_index(species)

    categories = []
    loc_map = {}  # sheet -> [(boss name-row, trainerId)] for hyperlink resolution
    total, matched, unmatched = 0, 0, []
    for sh in SHEETS:
        if sh not in wb.sheetnames:
            continue
        bosses_raw = extract_bosses(wb[sh], sidx, wb_f[sh])
        out = []
        for b in bosses_raw:
            total += 1
            tid, score = match_trainer(b, trainers)
            if tid is None:
                unmatched.append(sh + ' :: ' + b['name'] + '  ' + str(b['species_names']))
                continue
            matched += 1
            tr = trainers[str(tid)]
            lvls = [m['level'] for m in tr['hardcore']]
            out.append({'name': b['name'].title(), 'trainerName': tr['name'], 'trainerId': tid,
                        'minLevel': min(lvls), 'maxLevel': max(lvls), 'sprite': b.get('sprite', '')})
            loc_map.setdefault(sh, []).append((b['row'], tid))
        if out:
            categories.append({'name': sh, 'bosses': out})
        print('  %-16s %d bosses' % (sh, len(out)))
    for sh in loc_map:
        loc_map[sh].sort()

    info = build_info(wb['Main'])
    trainer_order = parse_trainer_order(wb['Trainer Order'], trainers, loc_map) if 'Trainer Order' in wb.sheetnames else []
    to_linked = sum(1 for e in trainer_order if e['trainerId'])
    out = {'info': info, 'categories': categories, 'trainerOrder': trainer_order}
    with open(os.path.join(DATA, 'hardcore.json'), 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)

    print('\nmatched %d / %d bosses across %d categories' % (matched, total, len(categories)))
    print('trainer order: %d entries, %d linked to a team' % (len(trainer_order), to_linked))
    print('info lines:', len(info))
    if unmatched:
        print('\nUNMATCHED (%d):' % len(unmatched))
        for u in unmatched[:40]:
            print('  ', u.encode('ascii', 'replace').decode())

if __name__ == '__main__':
    main()
